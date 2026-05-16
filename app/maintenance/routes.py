from datetime import datetime, date, time, timedelta
from io import BytesIO

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    send_file,
    jsonify,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.auth.routes import login_required, role_required
from app.extensions import db
from app.models import MaintenanceTicket, Store

maintenance_bp = Blueprint("maintenance", __name__, url_prefix="/maintenance")


def get_current_role():
    return (session.get("user_role") or "").strip()


def get_visible_stores():
    role = session.get("user_role")
    user_area = session.get("user_area")
    user_store = session.get("user_store")

    if role in ["admin", "maintenance"]:
        return Store.query.filter_by(is_active=True).order_by(Store.store_number.asc()).all()

    if role == "supervisor":
        return Store.query.filter_by(
            area_name=user_area,
            is_active=True
        ).order_by(Store.store_number.asc()).all()

    if role == "manager":
        return Store.query.filter_by(
            store_number=user_store,
            is_active=True
        ).order_by(Store.store_number.asc()).all()

    return []


def get_visible_store_numbers():
    return {store.store_number for store in get_visible_stores()}


def split_lines_to_tasks(text: str):
    return [line.strip() for line in (text or "").splitlines() if line.strip()]


def parse_date_field(value):
    value = (value or "").strip()
    if not value:
        return None

    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_time_field(value):
    value = (value or "").strip()
    if not value:
        return None

    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return None


def parse_int_field(value):
    value = (value or "").strip()
    if not value:
        return None

    try:
        parsed = int(value)
        return parsed if parsed >= 0 else None
    except ValueError:
        return None


def format_ticket_datetime(ticket):
    if not getattr(ticket, "scheduled_date", None):
        return ""

    date_part = ticket.scheduled_date.strftime("%b %d, %Y")

    if getattr(ticket, "scheduled_time", None):
        time_part = ticket.scheduled_time.strftime("%I:%M %p").lstrip("0")
        return f"{date_part} {time_part}"

    return date_part


def format_time(ticket):
    if not getattr(ticket, "scheduled_time", None):
        return "Any time"

    return ticket.scheduled_time.strftime("%I:%M %p").lstrip("0")


def build_time_slots():
    slots = []
    start_hour = 8
    end_hour = 20

    current = datetime.combine(date.today(), time(start_hour, 0))
    end = datetime.combine(date.today(), time(end_hour, 0))

    while current < end:
        slot_time = current.time()
        slots.append({
            "value": slot_time.strftime("%H:%M"),
            "label": slot_time.strftime("%I:%M %p").lstrip("0"),
            "time": slot_time,
        })
        current += timedelta(minutes=30)

    return slots


def get_filtered_tickets():
    visible_stores = get_visible_stores()
    visible_store_numbers = {store.store_number for store in visible_stores}

    status_filter = request.args.get("status", "").strip()
    store_filter = request.args.get("store", "").strip()

    tickets = MaintenanceTicket.query.order_by(
        MaintenanceTicket.created_at.asc(),
        MaintenanceTicket.id.asc()
    ).all()

    tickets = [t for t in tickets if t.store_number in visible_store_numbers]

    if status_filter:
        tickets = [t for t in tickets if t.status == status_filter]

    if store_filter:
        tickets = [t for t in tickets if t.store_number == store_filter]

    return tickets, visible_stores, status_filter, store_filter


def autosize_worksheet_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            except Exception:
                pass

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def style_header_row(worksheet, row_number=1):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_maintenance_excel(tickets, status_filter, store_filter):
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["Selected Store", store_filter or "All"])
    summary_ws.append(["Selected Status", status_filter or "All"])
    summary_ws.append(["Total Tickets", len(tickets)])

    style_header_row(summary_ws)
    autosize_worksheet_columns(summary_ws)

    tickets_ws = wb.create_sheet(title="Maintenance Tickets")
    tickets_ws.append([
        "Store",
        "Title",
        "Details",
        "Status",
        "Priority",
        "Assigned To",
        "Scheduled",
        "Estimated Minutes",
        "Source Type",
        "Created At",
        "SVR Report ID",
    ])

    for ticket in tickets:
        created_at_display = (
            ticket.created_at.strftime("%Y-%m-%d %I:%M %p")
            if ticket.created_at else ""
        )

        tickets_ws.append([
            ticket.store_number,
            ticket.title or "",
            ticket.details or "",
            ticket.status.replace("_", " ").title() if ticket.status else "",
            (getattr(ticket, "priority", None) or "normal").title(),
            getattr(ticket, "assigned_to", None) or "",
            format_ticket_datetime(ticket),
            getattr(ticket, "estimated_minutes", None) if getattr(ticket, "estimated_minutes", None) is not None else "",
            (ticket.source_type or "").upper(),
            created_at_display,
            ticket.svr_report_id if ticket.svr_report_id is not None else "",
        ])

    style_header_row(tickets_ws)
    autosize_worksheet_columns(tickets_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def apply_ticket_form_fields(ticket):
    ticket.store_number = request.form.get("store_number", "").strip()
    ticket.title = request.form.get("title", "").strip()
    ticket.details = request.form.get("details", "").strip()
    ticket.status = request.form.get("status", "").strip()

    ticket.assigned_to = request.form.get("assigned_to", "").strip() or None
    ticket.scheduled_date = parse_date_field(request.form.get("scheduled_date"))
    ticket.scheduled_time = parse_time_field(request.form.get("scheduled_time"))
    ticket.estimated_minutes = parse_int_field(request.form.get("estimated_minutes"))

    priority = request.form.get("priority", "normal").strip()
    valid_priorities = {"low", "normal", "high", "urgent"}
    ticket.priority = priority if priority in valid_priorities else "normal"


@maintenance_bp.route("/", methods=["GET", "POST"])
@login_required
@role_required("admin", "supervisor", "maintenance", "manager")
def index():
    visible_stores = get_visible_stores()
    visible_store_numbers = get_visible_store_numbers()
    role = get_current_role()

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        if action == "create":
            store_number = request.form.get("store_number", "").strip()
            title = request.form.get("title", "").strip()
            details = request.form.get("details", "").strip()

            assigned_to = request.form.get("assigned_to", "").strip() or None
            scheduled_date = parse_date_field(request.form.get("scheduled_date"))
            scheduled_time = parse_time_field(request.form.get("scheduled_time"))
            estimated_minutes = parse_int_field(request.form.get("estimated_minutes"))
            priority = request.form.get("priority", "normal").strip()

            valid_priorities = {"low", "normal", "high", "urgent"}
            if priority not in valid_priorities:
                priority = "normal"

            if store_number not in visible_store_numbers:
                flash("Invalid store selection.", "error")
                return redirect(url_for("maintenance.index"))

            title_lines = split_lines_to_tasks(title)

            if title_lines:
                for line in title_lines:
                    ticket = MaintenanceTicket(
                        store_number=store_number,
                        title=line,
                        details=details,
                        source_type="manual",
                        status="open",
                        assigned_to=assigned_to,
                        scheduled_date=scheduled_date,
                        scheduled_time=scheduled_time,
                        estimated_minutes=estimated_minutes,
                        priority=priority,
                    )
                    db.session.add(ticket)

                db.session.commit()

                if len(title_lines) == 1:
                    flash("Maintenance task created.", "success")
                else:
                    flash(f"{len(title_lines)} maintenance tasks created.", "success")

                return redirect(url_for("maintenance.index"))

            if not details:
                flash("Task title is required.", "error")
                return redirect(url_for("maintenance.index"))

            ticket = MaintenanceTicket(
                store_number=store_number,
                title="General maintenance task",
                details=details,
                source_type="manual",
                status="open",
                assigned_to=assigned_to,
                scheduled_date=scheduled_date,
                scheduled_time=scheduled_time,
                estimated_minutes=estimated_minutes,
                priority=priority,
            )
            db.session.add(ticket)
            db.session.commit()

            flash("Maintenance task created.", "success")
            return redirect(url_for("maintenance.index"))

        if action == "update":
            ticket_id = request.form.get("ticket_id", "").strip()
            ticket = MaintenanceTicket.query.get(ticket_id)

            if not ticket:
                flash("Ticket not found.", "error")
                return redirect(url_for("maintenance.index"))

            if ticket.store_number not in visible_store_numbers:
                flash("You do not have access to that ticket.", "error")
                return redirect(url_for("maintenance.index"))

            store_number = request.form.get("store_number", "").strip()
            title = request.form.get("title", "").strip()
            status = request.form.get("status", "").strip()

            valid_statuses = {"open", "assigned", "in_progress", "complete"}

            if store_number not in visible_store_numbers:
                flash("Invalid store selection.", "error")
                return redirect(url_for("maintenance.index"))

            if not title:
                flash("Task title is required.", "error")
                return redirect(url_for("maintenance.index"))

            if status not in valid_statuses:
                flash("Invalid status.", "error")
                return redirect(url_for("maintenance.index"))

            if role == "manager" and store_number != session.get("user_store"):
                flash("Managers can only manage tickets for their own store.", "error")
                return redirect(url_for("maintenance.index"))

            apply_ticket_form_fields(ticket)

            db.session.commit()
            flash("Maintenance task updated.", "success")
            return redirect(url_for("maintenance.index"))

        if action == "delete":
            if role == "manager":
                flash("Managers cannot delete maintenance tasks.", "error")
                return redirect(url_for("maintenance.index"))

            ticket_id = request.form.get("ticket_id", "").strip()
            ticket = MaintenanceTicket.query.get(ticket_id)

            if not ticket:
                flash("Ticket not found.", "error")
                return redirect(url_for("maintenance.index"))

            if ticket.store_number not in visible_store_numbers:
                flash("You do not have access to that ticket.", "error")
                return redirect(url_for("maintenance.index"))

            db.session.delete(ticket)
            db.session.commit()

            flash("Maintenance task deleted.", "success")
            return redirect(url_for("maintenance.index"))

    status_filter = request.args.get("status", "").strip()
    store_filter = request.args.get("store", "").strip()

    tickets = MaintenanceTicket.query.order_by(
        MaintenanceTicket.created_at.asc(),
        MaintenanceTicket.id.asc()
    ).all()

    tickets = [t for t in tickets if t.store_number in visible_store_numbers]

    if status_filter:
        tickets = [t for t in tickets if t.status == status_filter]

    if store_filter:
        tickets = [t for t in tickets if t.store_number == store_filter]

    open_tickets = sorted(
        [t for t in tickets if t.status == "open"],
        key=lambda t: (t.created_at or datetime.min, t.id or 0)
    )

    assigned_tickets = sorted(
        [t for t in tickets if t.status == "assigned"],
        key=lambda t: (t.created_at or datetime.min, t.id or 0)
    )

    in_progress_tickets = sorted(
        [t for t in tickets if t.status == "in_progress"],
        key=lambda t: (t.created_at or datetime.min, t.id or 0),
        reverse=True
    )

    complete_tickets = sorted(
        [t for t in tickets if t.status == "complete"],
        key=lambda t: (t.created_at or datetime.min, t.id or 0),
        reverse=True
    )

    return render_template(
        "maintenance.html",
        tickets=tickets,
        open_tickets=open_tickets,
        assigned_tickets=assigned_tickets,
        in_progress_tickets=in_progress_tickets,
        complete_tickets=complete_tickets,
        stores=visible_stores,
        status_filter=status_filter,
        store_filter=store_filter,
        user_role=role,
        format_ticket_datetime=format_ticket_datetime,
    )


@maintenance_bp.route("/calendar", methods=["GET", "POST"])
@login_required
@role_required("admin", "supervisor", "maintenance", "manager")
def calendar():
    visible_stores = get_visible_stores()
    visible_store_numbers = get_visible_store_numbers()
    role = get_current_role()

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        if action == "schedule":
            ticket_id = request.form.get("ticket_id", "").strip()
            ticket = MaintenanceTicket.query.get(ticket_id)

            if not ticket:
                flash("Ticket not found.", "error")
                return redirect(url_for("maintenance.calendar"))

            if ticket.store_number not in visible_store_numbers:
                flash("You do not have access to that ticket.", "error")
                return redirect(url_for("maintenance.calendar"))

            store_number = request.form.get("store_number", "").strip()
            title = request.form.get("title", "").strip()
            status = request.form.get("status", "").strip()

            valid_statuses = {"open", "assigned", "in_progress", "complete"}

            if store_number not in visible_store_numbers:
                flash("Invalid store selection.", "error")
                return redirect(url_for("maintenance.calendar"))

            if not title:
                flash("Task title is required.", "error")
                return redirect(url_for("maintenance.calendar"))

            if status not in valid_statuses:
                flash("Invalid status.", "error")
                return redirect(url_for("maintenance.calendar"))

            if role == "manager" and store_number != session.get("user_store"):
                flash("Managers can only manage tickets for their own store.", "error")
                return redirect(url_for("maintenance.calendar"))

            apply_ticket_form_fields(ticket)

            db.session.commit()
            flash("Maintenance schedule updated.", "success")

            start = request.form.get("calendar_start", "").strip()
            if start:
                return redirect(url_for("maintenance.calendar", start=start))

            return redirect(url_for("maintenance.calendar"))

    today = date.today()
    start_param = request.args.get("start", "").strip()
    week_start = parse_date_field(start_param)

    if not week_start:
        week_start = today - timedelta(days=today.weekday())

    week_end = week_start + timedelta(days=6)
    previous_week = week_start - timedelta(days=7)
    next_week = week_start + timedelta(days=7)
    time_slots = build_time_slots()

    tickets = MaintenanceTicket.query.order_by(
        MaintenanceTicket.scheduled_date.asc(),
        MaintenanceTicket.scheduled_time.asc(),
        MaintenanceTicket.created_at.asc(),
        MaintenanceTicket.id.asc()
    ).all()

    tickets = [t for t in tickets if t.store_number in visible_store_numbers]

    scheduled_tickets = [
        t for t in tickets
        if t.scheduled_date and week_start <= t.scheduled_date <= week_end
    ]

    unscheduled_tickets = [
        t for t in tickets
        if not t.scheduled_date and t.status != "complete"
    ]

    days = []
    for offset in range(7):
        current_day = week_start + timedelta(days=offset)

        day_tickets = [
            t for t in scheduled_tickets
            if t.scheduled_date == current_day
        ]

        any_time_tickets = sorted(
            [t for t in day_tickets if not t.scheduled_time],
            key=lambda t: (t.store_number or "", t.id or 0)
        )

        day_slots = []
        for slot in time_slots:
            slot_tickets = [
                t for t in day_tickets
                if t.scheduled_time and t.scheduled_time.strftime("%H:%M") == slot["value"]
            ]

            slot_tickets = sorted(
                slot_tickets,
                key=lambda t: (t.store_number or "", t.id or 0)
            )

            day_slots.append({
                "value": slot["value"],
                "label": slot["label"],
                "tickets": slot_tickets,
            })

        days.append({
            "date": current_day,
            "tickets": day_tickets,
            "any_time_tickets": any_time_tickets,
            "slots": day_slots,
        })

    return render_template(
        "maintenance_calendar.html",
        days=days,
        unscheduled_tickets=unscheduled_tickets,
        stores=visible_stores,
        week_start=week_start,
        week_end=week_end,
        previous_week=previous_week,
        next_week=next_week,
        user_role=role,
        format_ticket_datetime=format_ticket_datetime,
        format_time=format_time,
    )


@maintenance_bp.route("/calendar/move", methods=["POST"])
@login_required
@role_required("admin", "supervisor", "maintenance", "manager")
def move_calendar_ticket():
    visible_store_numbers = get_visible_store_numbers()

    ticket_id = request.form.get("ticket_id", "").strip()
    scheduled_date_raw = request.form.get("scheduled_date", "").strip()
    scheduled_time_raw = request.form.get("scheduled_time", "").strip()

    ticket = MaintenanceTicket.query.get(ticket_id)

    if not ticket:
        return jsonify({"ok": False, "message": "Ticket not found."}), 404

    if ticket.store_number not in visible_store_numbers:
        return jsonify({"ok": False, "message": "You do not have access to that ticket."}), 403

    if scheduled_date_raw == "unscheduled":
        ticket.scheduled_date = None
        ticket.scheduled_time = None
    else:
        scheduled_date = parse_date_field(scheduled_date_raw)

        if not scheduled_date:
            return jsonify({"ok": False, "message": "Invalid scheduled date."}), 400

        ticket.scheduled_date = scheduled_date

        if scheduled_time_raw:
            scheduled_time = parse_time_field(scheduled_time_raw)

            if not scheduled_time:
                return jsonify({"ok": False, "message": "Invalid scheduled time."}), 400

            ticket.scheduled_time = scheduled_time
        else:
            ticket.scheduled_time = None

    db.session.commit()

    return jsonify({
        "ok": True,
        "message": "Maintenance task moved.",
        "ticket_id": ticket.id,
    })


@maintenance_bp.route("/export/excel")
@login_required
@role_required("admin", "supervisor", "maintenance", "manager")
def export_excel():
    tickets, _, status_filter, store_filter = get_filtered_tickets()
    workbook_stream = create_maintenance_excel(tickets, status_filter, store_filter)

    filename_parts = ["maintenance_export"]
    if store_filter:
        filename_parts.append(store_filter)
    if status_filter:
        filename_parts.append(status_filter)

    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        workbook_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )