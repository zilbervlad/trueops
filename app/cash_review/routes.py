from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO

from flask import Blueprint, render_template, request, session, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.auth.routes import login_required, role_required
from app.models import CashLog, Store

cash_review_bp = Blueprint("cash_review", __name__, url_prefix="/cash-review")


def get_visible_stores():
    role = session.get("user_role")
    user_area = session.get("user_area")

    if role == "admin":
        return Store.query.filter_by(is_active=True).order_by(Store.store_number.asc()).all()

    if role == "supervisor":
        return Store.query.filter_by(
            area_name=user_area,
            is_active=True
        ).order_by(Store.store_number.asc()).all()

    return []


def build_closing_to_opening_diffs(logs):
    by_store = defaultdict(list)

    for log in logs:
        by_store[log.store_number].append(log)

    diff_rows = []

    for store_number, store_logs in by_store.items():
        ordered = sorted(
            store_logs,
            key=lambda x: (x.log_date, x.created_at or datetime.min, x.id or 0)
        )

        for i, current_log in enumerate(ordered):
            if current_log.shift_type != "closing":
                continue

            next_opening = None
            for future_log in ordered[i + 1:]:
                if future_log.shift_type == "opening":
                    next_opening = future_log
                    break

            if not next_opening:
                continue

            closing_total = current_log.total_cash or 0
            opening_total = next_opening.total_cash or 0
            diff_amount = opening_total - closing_total

            diff_rows.append({
                "store_number": store_number,
                "closing_date": current_log.log_date,
                "opening_date": next_opening.log_date,
                "closing_total": closing_total,
                "opening_total": opening_total,
                "difference": diff_amount,
                "closing_manager": current_log.manager_name,
                "opening_manager": next_opening.manager_name,
            })

    diff_rows.sort(
        key=lambda x: (x["closing_date"], x["store_number"]),
        reverse=True
    )
    return diff_rows


def build_cash_review_payload():
    visible_stores = get_visible_stores()
    visible_store_numbers = {store.store_number for store in visible_stores}

    store_filter = (request.args.get("store") or "").strip()
    shift_filter = (request.args.get("shift") or "").strip()
    date_filter = (request.args.get("date") or "").strip()

    today = datetime.today().date()
    selected_date = None

    if date_filter:
        try:
            selected_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
        except ValueError:
            selected_date = None
            date_filter = ""

    dashboard_date = selected_date or today

    query = CashLog.query.filter(
        CashLog.store_number.in_(visible_store_numbers),
        CashLog.log_date == dashboard_date
    ).order_by(
        CashLog.created_at.desc()
    )

    if store_filter:
        query = query.filter(CashLog.store_number == store_filter)

    if shift_filter:
        query = query.filter(CashLog.shift_type == shift_filter)

    logs = query.limit(100).all()

    midshift_logs = [
        log for log in logs
        if log.shift_type == "midshift"
    ]

    midshift_logs = sorted(
        midshift_logs,
        key=lambda x: (
            abs(x.cash_over_short or 0),
            x.log_date,
            x.store_number
        ),
        reverse=True
    )

    diff_base_query = CashLog.query.filter(
        CashLog.store_number.in_(visible_store_numbers)
    )

    if store_filter:
        diff_base_query = diff_base_query.filter(CashLog.store_number == store_filter)

    diff_base_query = diff_base_query.filter(
        CashLog.log_date >= dashboard_date - timedelta(days=1),
        CashLog.log_date <= dashboard_date
    )

    diff_logs = diff_base_query.all()
    closing_opening_diffs = build_closing_to_opening_diffs(diff_logs)

    summary = {
        "log_count": len(logs),
        "midshift_count": len(midshift_logs),
        "diff_pair_count": len(closing_opening_diffs),
        "stores_in_scope": len(visible_stores),
    }

    return {
        "stores": visible_stores,
        "logs": logs,
        "midshift_logs": midshift_logs,
        "closing_opening_diffs": closing_opening_diffs,
        "store_filter": store_filter,
        "shift_filter": shift_filter,
        "date_filter": date_filter,
        "summary": summary,
    }


def autosize_worksheet_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            except Exception:
                pass

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def style_header_row(worksheet, row_number=1):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_cash_review_excel(payload):
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["Metric", "Value"])

    summary_rows = [
        ("Selected Store", payload["store_filter"] or "All"),
        ("Selected Shift", payload["shift_filter"] or "All"),
        ("Selected Date", payload["date_filter"] or "Today"),
        ("Stores In Scope", payload["summary"]["stores_in_scope"]),
        ("Recent Cash Logs", payload["summary"]["log_count"]),
        ("Midshift Exceptions", payload["summary"]["midshift_count"]),
        ("Closing to Opening Pairs", payload["summary"]["diff_pair_count"]),
    ]

    for row in summary_rows:
        summary_ws.append(row)

    style_header_row(summary_ws)
    autosize_worksheet_columns(summary_ws)

    logs_ws = wb.create_sheet(title="Recent Cash Logs")
    logs_ws.append([
        "Store",
        "Date",
        "Shift",
        "Back Till",
        "Front Till",
        "Driver Banks",
        "Total Cash",
        "Amount To Account For",
        "Cash Over / Short",
        "Manager",
    ])

    for log in payload["logs"]:
        logs_ws.append([
            log.store_number,
            log.log_date.strftime("%Y-%m-%d") if log.log_date else "",
            log.shift_type.title() if log.shift_type else "",
            log.back_till if log.back_till is not None else "",
            log.front_till if log.front_till is not None else "",
            log.driver_banks if log.driver_banks is not None else "",
            log.total_cash if log.total_cash is not None else "",
            log.amount_to_account_for if log.amount_to_account_for is not None else "",
            log.cash_over_short if log.cash_over_short is not None else "",
            log.manager_name or "",
        ])

    style_header_row(logs_ws)
    autosize_worksheet_columns(logs_ws)

    midshift_ws = wb.create_sheet(title="Midshift Exceptions")
    midshift_ws.append([
        "Store",
        "Date",
        "Total Cash",
        "Amount To Account For",
        "Cash Over / Short",
        "Manager",
    ])

    for log in payload["midshift_logs"]:
        midshift_ws.append([
            log.store_number,
            log.log_date.strftime("%Y-%m-%d") if log.log_date else "",
            log.total_cash if log.total_cash is not None else "",
            log.amount_to_account_for if log.amount_to_account_for is not None else "",
            log.cash_over_short if log.cash_over_short is not None else "",
            log.manager_name or "",
        ])

    style_header_row(midshift_ws)
    autosize_worksheet_columns(midshift_ws)

    diff_ws = wb.create_sheet(title="Closing Opening Diff")
    diff_ws.append([
        "Store",
        "Closing Date",
        "Closing Total",
        "Closing Manager",
        "Opening Date",
        "Opening Total",
        "Opening Manager",
        "Difference",
    ])

    for row in payload["closing_opening_diffs"]:
        diff_ws.append([
            row["store_number"],
            row["closing_date"].strftime("%Y-%m-%d") if row["closing_date"] else "",
            row["closing_total"],
            row["closing_manager"] or "",
            row["opening_date"].strftime("%Y-%m-%d") if row["opening_date"] else "",
            row["opening_total"],
            row["opening_manager"] or "",
            row["difference"],
        ])

    style_header_row(diff_ws)
    autosize_worksheet_columns(diff_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@cash_review_bp.route("/", methods=["GET"])
@login_required
@role_required("admin", "supervisor")
def index():
    payload = build_cash_review_payload()

    return render_template(
        "cash_review.html",
        stores=payload["stores"],
        logs=payload["logs"],
        midshift_logs=payload["midshift_logs"],
        closing_opening_diffs=payload["closing_opening_diffs"],
        store_filter=payload["store_filter"],
        shift_filter=payload["shift_filter"],
        date_filter=payload["date_filter"],
        summary=payload["summary"],
    )


@cash_review_bp.route("/export/excel", methods=["GET"])
@login_required
@role_required("admin", "supervisor")
def export_excel():
    payload = build_cash_review_payload()
    workbook_stream = create_cash_review_excel(payload)

    filename_parts = ["cash_review"]
    if payload["store_filter"]:
        filename_parts.append(payload["store_filter"])
    if payload["shift_filter"]:
        filename_parts.append(payload["shift_filter"])
    if payload["date_filter"]:
        filename_parts.append(payload["date_filter"])
    else:
        filename_parts.append("today")

    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        workbook_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )