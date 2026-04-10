from collections import defaultdict
from datetime import date, timedelta, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from flask import Blueprint, render_template, request, session, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.auth.routes import login_required, role_required
from app.models import Store, DailyChecklist, NightlyNumbersReport

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")

APP_TZ = ZoneInfo("America/New_York")
UTC_TZ = ZoneInfo("UTC")


def utc_naive_to_et(dt):
    if not dt:
        return None
    return dt.replace(tzinfo=UTC_TZ).astimezone(APP_TZ)


def get_visible_stores():
    role = session.get("user_role")
    user_area = session.get("user_area")
    user_store = session.get("user_store")

    if role == "admin":
        return Store.query.filter_by(is_active=True).order_by(
            Store.area_name.asc(),
            Store.store_number.asc()
        ).all()

    if role == "supervisor":
        return Store.query.filter_by(
            area_name=user_area,
            is_active=True
        ).order_by(
            Store.area_name.asc(),
            Store.store_number.asc()
        ).all()

    if role == "manager":
        return Store.query.filter_by(
            store_number=user_store,
            is_active=True
        ).order_by(Store.store_number.asc()).all()

    return []


def parse_report_dates():
    today = date.today()
    default_start = today - timedelta(days=6)
    default_end = today

    start_date_str = request.args.get("start_date", default_start.strftime("%Y-%m-%d")).strip()
    end_date_str = request.args.get("end_date", default_end.strftime("%Y-%m-%d")).strip()

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    except ValueError:
        start_date = default_start
        start_date_str = start_date.strftime("%Y-%m-%d")

    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        end_date = default_end
        end_date_str = end_date.strftime("%Y-%m-%d")

    if end_date < start_date:
        start_date, end_date = end_date, start_date
        start_date_str = start_date.strftime("%Y-%m-%d")
        end_date_str = end_date.strftime("%Y-%m-%d")

    return start_date, end_date, start_date_str, end_date_str


def calculate_section_percent(daily, section_name):
    if not daily or not getattr(daily, "items", None):
        return 0.0

    section_items = [item for item in daily.items if item.section_name == section_name]
    total = len(section_items)

    if total == 0:
        return 0.0

    completed = sum(1 for item in section_items if item.is_completed)
    return round((completed / total) * 100, 1)


def build_report_payload():
    visible_stores = get_visible_stores()

    start_date, end_date, start_date_str, end_date_str = parse_report_dates()

    q = request.args.get("q", "").strip()
    selected_store = request.args.get("store_number", "").strip()
    selected_area = request.args.get("area_name", "").strip()
    manager_name = request.args.get("manager_name", "").strip()
    nightly_filter = request.args.get("nightly_filter", "").strip()
    show_task_analysis = request.args.get("show_task_analysis", "").strip() == "1"

    area_options = sorted({store.area_name for store in visible_stores if store.area_name})

    filtered_stores = visible_stores

    if selected_store:
        filtered_stores = [
            store for store in filtered_stores
            if store.store_number == selected_store
        ]

    if selected_area:
        filtered_stores = [
            store for store in filtered_stores
            if store.area_name == selected_area
        ]

    if q:
        q_lower = q.lower()
        filtered_stores = [
            store for store in filtered_stores
            if (
                q_lower in (store.store_number or "").lower()
                or q_lower in (store.store_name or "").lower()
                or q_lower in (store.area_name or "").lower()
            )
        ]

    filtered_store_numbers = {store.store_number for store in filtered_stores}

    checklist_rows = []
    if filtered_store_numbers:
        checklist_rows = DailyChecklist.query.filter(
            DailyChecklist.store_number.in_(filtered_store_numbers),
            DailyChecklist.checklist_date >= start_date,
            DailyChecklist.checklist_date <= end_date,
        ).order_by(
            DailyChecklist.checklist_date.desc(),
            DailyChecklist.store_number.asc()
        ).all()

    nightly_rows = []
    if filtered_store_numbers:
        nightly_rows = NightlyNumbersReport.query.filter(
            NightlyNumbersReport.store_number.in_(filtered_store_numbers),
            NightlyNumbersReport.report_date >= start_date,
            NightlyNumbersReport.report_date <= end_date,
        ).order_by(
            NightlyNumbersReport.report_date.desc(),
            NightlyNumbersReport.store_number.asc()
        ).all()

    if manager_name:
        manager_lower = manager_name.lower()
        checklist_rows = [
            row for row in checklist_rows
            if (
                manager_lower in (row.manager_on_duty or "").lower()
                or manager_lower in (row.opening_manager or "").lower()
                or manager_lower in (row.closing_manager or "").lower()
            )
        ]

    if q:
        q_lower = q.lower()
        checklist_rows = [
            row for row in checklist_rows
            if (
                q_lower in (row.store_number or "").lower()
                or q_lower in (row.manager_on_duty or "").lower()
                or q_lower in (row.opening_manager or "").lower()
                or q_lower in (row.closing_manager or "").lower()
            )
        ]

    if nightly_filter == "missing":
        stores_with_nightly = {row.store_number for row in nightly_rows}
        filtered_stores = [
            store for store in filtered_stores
            if store.store_number not in stores_with_nightly
        ]
    elif nightly_filter == "adt_high":
        stores_with_high_adt = {
            row.store_number
            for row in nightly_rows
            if row.adt is not None and row.adt > 25
        }
        filtered_stores = [
            store for store in filtered_stores
            if store.store_number in stores_with_high_adt
        ]
    elif nightly_filter == "labor_high":
        stores_with_high_labor = {
            row.store_number
            for row in nightly_rows
            if (
                row.variable_labor is not None
                and row.labor_goal is not None
                and row.variable_labor > row.labor_goal
            )
        }
        filtered_stores = [
            store for store in filtered_stores
            if store.store_number in stores_with_high_labor
        ]
    elif nightly_filter == "submitted":
        stores_with_nightly = {row.store_number for row in nightly_rows}
        filtered_stores = [
            store for store in filtered_stores
            if store.store_number in stores_with_nightly
        ]

    filtered_store_numbers = {store.store_number for store in filtered_stores}

    checklist_rows = [
        row for row in checklist_rows
        if row.store_number in filtered_store_numbers
    ]

    nightly_rows = [
        row for row in nightly_rows
        if row.store_number in filtered_store_numbers
    ]

    checklist_store_numbers = {row.store_number for row in checklist_rows}

    if selected_store or selected_area or q or manager_name:
        filtered_stores = [
            store for store in filtered_stores
            if store.store_number in checklist_store_numbers
            or store.store_number in {row.store_number for row in nightly_rows}
            or (
                q
                and (
                    q.lower() in (store.store_number or "").lower()
                    or q.lower() in (store.store_name or "").lower()
                    or q.lower() in (store.area_name or "").lower()
                )
            )
            or nightly_filter == "missing"
        ]

    days_in_range = (end_date - start_date).days + 1

    summary = {
        "avg_completion": 0.0,
        "avg_integrity": 0.0,
        "completed_count": 0,
        "in_progress_count": 0,
        "not_started_count": 0,
        "days_in_range": days_in_range,
        "stores_in_scope": len(filtered_stores),
        "total_checklists": len(checklist_rows),
        "stores_with_nightly": len({row.store_number for row in nightly_rows}),
        "stores_missing_nightly": 0,
    }

    if checklist_rows:
        summary["avg_completion"] = round(
            sum(row.percent_complete for row in checklist_rows) / len(checklist_rows), 1
        )

        integrity_values = [row.integrity_score for row in checklist_rows if row.integrity_score > 0]
        summary["avg_integrity"] = round(
            sum(integrity_values) / len(integrity_values), 1
        ) if integrity_values else 0.0

        summary["completed_count"] = sum(1 for row in checklist_rows if row.status == "completed")
        summary["in_progress_count"] = sum(1 for row in checklist_rows if row.status == "in_progress")

    possible_checklists = len(filtered_stores) * days_in_range
    summary["not_started_count"] = max(possible_checklists - len(checklist_rows), 0)
    summary["stores_missing_nightly"] = max(
        len(filtered_stores) - len({row.store_number for row in nightly_rows}),
        0
    )

    store_report_rows = []

    for store in filtered_stores:
        store_rows = [row for row in checklist_rows if row.store_number == store.store_number]
        store_nightly_rows = [row for row in nightly_rows if row.store_number == store.store_number]

        latest_nightly = store_nightly_rows[0] if store_nightly_rows else None

        checklist_count = len(store_rows)
        avg_completion = round(
            sum(row.percent_complete for row in store_rows) / checklist_count, 1
        ) if checklist_count else 0.0

        integrity_values = [row.integrity_score for row in store_rows if row.integrity_score > 0]
        avg_integrity = round(
            sum(integrity_values) / len(integrity_values), 1
        ) if integrity_values else 0.0

        avg_opening = round(
            sum(calculate_section_percent(row, "Before Open / Before 10:30") for row in store_rows) / checklist_count, 1
        ) if checklist_count else 0.0

        avg_restock = round(
            sum(calculate_section_percent(row, "3-O'Clock Restock") for row in store_rows) / checklist_count, 1
        ) if checklist_count else 0.0

        avg_manager_walk = round(
            sum(calculate_section_percent(row, "Manager's Walk") for row in store_rows) / checklist_count, 1
        ) if checklist_count else 0.0

        manager_names = sorted({
            name.strip()
            for row in store_rows
            for name in [row.manager_on_duty, row.opening_manager, row.closing_manager]
            if name and name.strip()
        })

        completed_count = sum(1 for row in store_rows if row.status == "completed")
        in_progress_count = sum(1 for row in store_rows if row.status == "in_progress")
        not_started_count = max(days_in_range - checklist_count, 0)

        store_report_rows.append({
            "store_number": store.store_number,
            "store_name": store.store_name or f"Store {store.store_number}",
            "area_name": store.area_name or "Unassigned",
            "avg_completion": avg_completion,
            "avg_integrity": avg_integrity,
            "avg_opening": avg_opening,
            "avg_restock": avg_restock,
            "avg_manager_walk": avg_manager_walk,
            "completed_count": completed_count,
            "in_progress_count": in_progress_count,
            "not_started_count": not_started_count,
            "checklist_count": checklist_count,
            "manager_names": manager_names,
            "has_nightly": latest_nightly is not None,
            "nightly_date": latest_nightly.report_date if latest_nightly else None,
            "nightly_adt": latest_nightly.adt if latest_nightly else None,
            "nightly_labor": latest_nightly.variable_labor if latest_nightly else None,
            "nightly_labor_goal": latest_nightly.labor_goal if latest_nightly else None,
        })

    store_report_rows = sorted(
        store_report_rows,
        key=lambda x: (x["area_name"], x["store_number"])
    )

    area_rollups = defaultdict(lambda: {
        "store_count": 0,
        "avg_completion_total": 0.0,
        "avg_integrity_total": 0.0,
        "avg_opening_total": 0.0,
        "avg_restock_total": 0.0,
        "avg_manager_walk_total": 0.0,
        "completed_count": 0,
        "in_progress_count": 0,
        "not_started_count": 0,
        "checklist_count": 0,
        "stores_with_nightly": 0,
    })

    for row in store_report_rows:
        rollup = area_rollups[row["area_name"]]
        rollup["store_count"] += 1
        rollup["avg_completion_total"] += row["avg_completion"]
        rollup["avg_integrity_total"] += row["avg_integrity"]
        rollup["avg_opening_total"] += row["avg_opening"]
        rollup["avg_restock_total"] += row["avg_restock"]
        rollup["avg_manager_walk_total"] += row["avg_manager_walk"]
        rollup["completed_count"] += row["completed_count"]
        rollup["in_progress_count"] += row["in_progress_count"]
        rollup["not_started_count"] += row["not_started_count"]
        rollup["checklist_count"] += row["checklist_count"]
        rollup["stores_with_nightly"] += 1 if row["has_nightly"] else 0

    area_report_rows = []
    for area_name, rollup in sorted(area_rollups.items()):
        store_count = rollup["store_count"]
        area_report_rows.append({
            "area_name": area_name,
            "store_count": store_count,
            "avg_completion": round(rollup["avg_completion_total"] / store_count, 1) if store_count else 0.0,
            "avg_integrity": round(rollup["avg_integrity_total"] / store_count, 1) if store_count else 0.0,
            "avg_opening": round(rollup["avg_opening_total"] / store_count, 1) if store_count else 0.0,
            "avg_restock": round(rollup["avg_restock_total"] / store_count, 1) if store_count else 0.0,
            "avg_manager_walk": round(rollup["avg_manager_walk_total"] / store_count, 1) if store_count else 0.0,
            "completed_count": rollup["completed_count"],
            "in_progress_count": rollup["in_progress_count"],
            "not_started_count": rollup["not_started_count"],
            "checklist_count": rollup["checklist_count"],
            "stores_with_nightly": rollup["stores_with_nightly"],
            "stores_missing_nightly": max(store_count - rollup["stores_with_nightly"], 0),
        })

    task_analysis = None
    if show_task_analysis:
        ranked_stores = sorted(
            [
                row for row in store_report_rows
                if row["checklist_count"] > 0
            ],
            key=lambda x: (
                -x["avg_completion"],
                -x["avg_integrity"],
                -x["avg_opening"],
                x["store_number"],
            )
        )

        lowest_integrity = sorted(
            [
                row for row in store_report_rows
                if row["checklist_count"] > 0
            ],
            key=lambda x: (
                x["avg_integrity"],
                x["avg_completion"],
                x["store_number"],
            )
        )

        completion_gaps = sorted(
            [
                row for row in store_report_rows
                if row["avg_restock"] < 100 or row["avg_manager_walk"] < 100 or row["avg_opening"] < 100
            ],
            key=lambda x: (
                x["avg_opening"],
                x["avg_restock"],
                x["avg_manager_walk"],
                x["store_number"],
            )
        )

        task_analysis = {
            "top_ranked_stores": ranked_stores[:5],
            "lowest_integrity_stores": lowest_integrity[:5],
            "biggest_completion_gaps": completion_gaps[:5],
        }

    return {
        "stores": visible_stores,
        "area_options": area_options,
        "selected_store": selected_store,
        "selected_area": selected_area,
        "manager_name": manager_name,
        "nightly_filter": nightly_filter,
        "q": q,
        "start_date": start_date_str,
        "end_date": end_date_str,
        "show_task_analysis": show_task_analysis,
        "summary": summary,
        "store_report_rows": store_report_rows,
        "area_report_rows": area_report_rows,
        "task_analysis": task_analysis,
        "checklist_rows": checklist_rows,
    }


def autosize_worksheet_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            except Exception:
                pass

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def style_header_row(worksheet, row_number=1):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_excel_report(payload):
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"

    summary_ws.append(["Metric", "Value"])
    summary = payload["summary"]
    summary_rows = [
        ("Start Date", payload["start_date"]),
        ("End Date", payload["end_date"]),
        ("Search Query", payload["q"] or "All"),
        ("Selected Store", payload["selected_store"] or "All"),
        ("Selected Area", payload["selected_area"] or "All"),
        ("Manager Filter", payload["manager_name"] or "All"),
        ("Nightly Filter", payload["nightly_filter"] or "All"),
        ("Days In Range", summary["days_in_range"]),
        ("Stores In Scope", summary["stores_in_scope"]),
        ("Total Checklists", summary["total_checklists"]),
        ("Avg Completion %", summary["avg_completion"]),
        ("Avg Integrity %", summary["avg_integrity"]),
        ("Completed Count", summary["completed_count"]),
        ("In Progress Count", summary["in_progress_count"]),
        ("Not Started Count", summary["not_started_count"]),
        ("Stores With Nightly Numbers", summary["stores_with_nightly"]),
        ("Stores Missing Nightly Numbers", summary["stores_missing_nightly"]),
    ]
    for row in summary_rows:
        summary_ws.append(row)

    style_header_row(summary_ws)
    autosize_worksheet_columns(summary_ws)

    store_ws = wb.create_sheet(title="Store Performance")
    store_ws.append([
        "Store Number",
        "Store Name",
        "Area",
        "Managers",
        "Avg Completion %",
        "Avg Integrity %",
        "Opening %",
        "3 O'Clock Restock %",
        "Manager's Walk %",
        "Has Nightly Numbers",
        "Nightly Date",
        "Nightly ADT",
        "Nightly Labor",
        "Nightly Labor Goal",
        "Total Checklists",
    ])

    for row in payload["store_report_rows"]:
        store_ws.append([
            row["store_number"],
            row["store_name"],
            row["area_name"],
            ", ".join(row["manager_names"]) if row["manager_names"] else "—",
            row["avg_completion"],
            row["avg_integrity"],
            row["avg_opening"],
            row["avg_restock"],
            row["avg_manager_walk"],
            "Yes" if row["has_nightly"] else "No",
            row["nightly_date"].strftime("%Y-%m-%d") if row["nightly_date"] else "",
            row["nightly_adt"] if row["nightly_adt"] is not None else "",
            row["nightly_labor"] if row["nightly_labor"] is not None else "",
            row["nightly_labor_goal"] if row["nightly_labor_goal"] is not None else "",
            row["checklist_count"],
        ])

    style_header_row(store_ws)
    autosize_worksheet_columns(store_ws)

    area_ws = wb.create_sheet(title="Area Performance")
    area_ws.append([
        "Area",
        "Store Count",
        "Avg Completion %",
        "Avg Integrity %",
        "Avg Opening %",
        "Avg 3 O'Clock Restock %",
        "Avg Manager's Walk %",
        "Stores With Nightly Numbers",
        "Stores Missing Nightly Numbers",
        "Total Checklists",
    ])

    for row in payload["area_report_rows"]:
        area_ws.append([
            row["area_name"],
            row["store_count"],
            row["avg_completion"],
            row["avg_integrity"],
            row["avg_opening"],
            row["avg_restock"],
            row["avg_manager_walk"],
            row["stores_with_nightly"],
            row["stores_missing_nightly"],
            row["checklist_count"],
        ])

    style_header_row(area_ws)
    autosize_worksheet_columns(area_ws)

    timeline_ws = wb.create_sheet(title="Timeline Audit")
    timeline_ws.append([
        "Store",
        "Checklist Date",
        "Task",
        "Section",
        "Completed Time (ET)",
        "Gap From Previous (min)",
    ])

    filtered_lookup = {
        (row.store_number, row.checklist_date): row
        for row in payload["checklist_rows"]
    }

    for _, checklist in sorted(filtered_lookup.items(), key=lambda x: (x[0][0], x[0][1])):
        if not getattr(checklist, "items", None):
            continue

        completed_items = [
            item for item in checklist.items
            if item.is_completed and item.completed_at
        ]
        completed_items = sorted(completed_items, key=lambda x: x.completed_at)

        prev_time = None

        for item in completed_items:
            completed_at_et = utc_naive_to_et(item.completed_at)

            gap_minutes = None
            if prev_time:
                gap_minutes = round(
                    (item.completed_at - prev_time).total_seconds() / 60, 1
                )

            timeline_ws.append([
                checklist.store_number,
                checklist.checklist_date.strftime("%Y-%m-%d"),
                item.task_text,
                item.section_name,
                completed_at_et.strftime("%I:%M %p") if completed_at_et else "",
                gap_minutes if gap_minutes is not None else "",
            ])

            prev_time = item.completed_at

    style_header_row(timeline_ws)
    autosize_worksheet_columns(timeline_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_store_detail_payload(store_number):
    visible_stores = get_visible_stores()
    visible_store_map = {store.store_number: store for store in visible_stores}

    store = visible_store_map.get(store_number)
    if not store:
        return None

    start_date, end_date, start_date_str, end_date_str = parse_report_dates()

    store_rows = DailyChecklist.query.filter(
        DailyChecklist.store_number == store_number,
        DailyChecklist.checklist_date >= start_date,
        DailyChecklist.checklist_date <= end_date,
    ).order_by(
        DailyChecklist.checklist_date.desc()
    ).all()

    checklist_count = len(store_rows)
    days_in_range = (end_date - start_date).days + 1

    avg_completion = round(
        sum(row.percent_complete for row in store_rows) / checklist_count, 1
    ) if checklist_count else 0.0

    integrity_values = [row.integrity_score for row in store_rows if row.integrity_score > 0]
    avg_integrity = round(
        sum(integrity_values) / len(integrity_values), 1
    ) if integrity_values else 0.0

    completed_count = sum(1 for row in store_rows if row.status == "completed")
    in_progress_count = sum(1 for row in store_rows if row.status == "in_progress")
    not_started_count = max(days_in_range - checklist_count, 0)

    manager_names = sorted({
        name.strip()
        for row in store_rows
        for name in [row.manager_on_duty, row.opening_manager, row.closing_manager]
        if name and name.strip()
    })

    daily_rows = []
    for row in store_rows:
        managers_for_day = [
            value for value in [
                row.manager_on_duty,
                row.opening_manager,
                row.closing_manager,
            ]
            if value and value.strip()
        ]

        timeline = []

        if getattr(row, "items", None):
            completed_items = [
                item for item in row.items
                if item.is_completed and item.completed_at
            ]

            completed_items = sorted(completed_items, key=lambda x: x.completed_at)

            prev_time = None

            for item in completed_items:
                completed_at_et = utc_naive_to_et(item.completed_at)
                gap_minutes = None

                if prev_time:
                    gap_minutes = round(
                        (item.completed_at - prev_time).total_seconds() / 60, 1
                    )

                timeline.append({
                    "task_text": item.task_text,
                    "section": item.section_name,
                    "completed_at": completed_at_et,
                    "completed_at_display": completed_at_et.strftime("%I:%M %p") if completed_at_et else "",
                    "gap_minutes": gap_minutes,
                })

                prev_time = item.completed_at

        daily_rows.append({
            "checklist_date": row.checklist_date,
            "status": row.status,
            "percent_complete": row.percent_complete,
            "integrity_score": row.integrity_score,
            "manager_display": " / ".join(dict.fromkeys(managers_for_day)) if managers_for_day else "—",
            "timeline": timeline,
        })

    summary = {
        "avg_completion": avg_completion,
        "avg_integrity": avg_integrity,
        "completed_count": completed_count,
        "in_progress_count": in_progress_count,
        "not_started_count": not_started_count,
        "checklist_count": checklist_count,
        "days_in_range": days_in_range,
    }

    back_args = {
        "start_date": start_date_str,
        "end_date": end_date_str,
    }

    q = request.args.get("q", "").strip()
    selected_area = request.args.get("area_name", "").strip()
    selected_store = request.args.get("store_number", "").strip()
    manager_name = request.args.get("manager_name", "").strip()
    nightly_filter = request.args.get("nightly_filter", "").strip()
    show_task_analysis = request.args.get("show_task_analysis", "").strip()

    if q:
        back_args["q"] = q
    if selected_area:
        back_args["area_name"] = selected_area
    if selected_store:
        back_args["store_number"] = selected_store
    if manager_name:
        back_args["manager_name"] = manager_name
    if nightly_filter:
        back_args["nightly_filter"] = nightly_filter
    if show_task_analysis == "1":
        back_args["show_task_analysis"] = "1"

    export_args = dict(back_args)

    return {
        "store": store,
        "store_rows": daily_rows,
        "summary": summary,
        "manager_names": manager_names,
        "start_date": start_date_str,
        "end_date": end_date_str,
        "back_args": back_args,
        "export_args": export_args,
    }


def create_store_detail_excel(payload):
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Store Summary"
    summary_ws.append(["Metric", "Value"])

    summary = payload["summary"]
    summary_rows = [
        ("Store Number", payload["store"].store_number),
        ("Store Name", payload["store"].store_name or f"Store {payload['store'].store_number}"),
        ("Area", payload["store"].area_name or ""),
        ("Start Date", payload["start_date"]),
        ("End Date", payload["end_date"]),
        ("Avg Completion %", summary["avg_completion"]),
        ("Avg Integrity %", summary["avg_integrity"]),
        ("Completed Count", summary["completed_count"]),
        ("In Progress Count", summary["in_progress_count"]),
        ("Not Started Count", summary["not_started_count"]),
        ("Checklist Count", summary["checklist_count"]),
    ]
    for row in summary_rows:
        summary_ws.append(row)

    style_header_row(summary_ws)
    autosize_worksheet_columns(summary_ws)

    history_ws = wb.create_sheet(title="Daily History")
    history_ws.append([
        "Checklist Date",
        "Status",
        "Completion %",
        "Integrity %",
        "Managers",
        "Timeline Events",
    ])

    for row in payload["store_rows"]:
        history_ws.append([
            row["checklist_date"].strftime("%Y-%m-%d"),
            row["status"].replace("_", " ").title(),
            row["percent_complete"],
            row["integrity_score"],
            row["manager_display"],
            len(row["timeline"]),
        ])

    style_header_row(history_ws)
    autosize_worksheet_columns(history_ws)

    timeline_ws = wb.create_sheet(title="Timeline Audit")
    timeline_ws.append([
        "Checklist Date",
        "Task",
        "Section",
        "Completed Time (ET)",
        "Gap From Previous (min)",
    ])

    for row in payload["store_rows"]:
        if not row["timeline"]:
            timeline_ws.append([
                row["checklist_date"].strftime("%Y-%m-%d"),
                "No completed tasks",
                "",
                "",
                "",
            ])
            continue

        for item in row["timeline"]:
            timeline_ws.append([
                row["checklist_date"].strftime("%Y-%m-%d"),
                item["task_text"],
                item["section"],
                item["completed_at_display"],
                item["gap_minutes"] if item["gap_minutes"] is not None else "",
            ])

    style_header_row(timeline_ws)
    autosize_worksheet_columns(timeline_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@reports_bp.route("/")
@login_required
@role_required("admin", "supervisor")
def index():
    payload = build_report_payload()

    return render_template(
        "reports.html",
        stores=payload["stores"],
        area_options=payload["area_options"],
        selected_store=payload["selected_store"],
        selected_area=payload["selected_area"],
        manager_name=payload["manager_name"],
        nightly_filter=payload["nightly_filter"],
        q=payload["q"],
        start_date=payload["start_date"],
        end_date=payload["end_date"],
        show_task_analysis=payload["show_task_analysis"],
        summary=payload["summary"],
        store_report_rows=payload["store_report_rows"],
        area_report_rows=payload["area_report_rows"],
        task_analysis=payload["task_analysis"],
    )


@reports_bp.route("/export/excel")
@login_required
@role_required("admin", "supervisor")
def export_excel():
    payload = build_report_payload()
    workbook_stream = create_excel_report(payload)

    filename = f"checklist_report_{payload['start_date']}_to_{payload['end_date']}.xlsx"

    return send_file(
        workbook_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reports_bp.route("/store/<store_number>")
@login_required
@role_required("admin", "supervisor")
def store_detail(store_number):
    payload = build_store_detail_payload(store_number)

    if not payload:
        return render_template(
            "store_report_detail.html",
            store=None,
            store_rows=[],
            summary=None,
            manager_names=[],
            start_date="",
            end_date="",
            back_args={},
            export_args={},
        )

    return render_template(
        "store_report_detail.html",
        store=payload["store"],
        store_rows=payload["store_rows"],
        summary=payload["summary"],
        manager_names=payload["manager_names"],
        start_date=payload["start_date"],
        end_date=payload["end_date"],
        back_args=payload["back_args"],
        export_args=payload["export_args"],
    )


@reports_bp.route("/store/<store_number>/export/excel")
@login_required
@role_required("admin", "supervisor")
def export_store_detail_excel(store_number):
    payload = build_store_detail_payload(store_number)

    if not payload:
        empty_stream = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Store Summary"
        ws.append(["Error", "Store not found or not visible"])
        wb.save(empty_stream)
        empty_stream.seek(0)
        return send_file(
            empty_stream,
            as_attachment=True,
            download_name="store_report_error.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    workbook_stream = create_store_detail_excel(payload)
    filename = f"store_{store_number}_timeline_{payload['start_date']}_to_{payload['end_date']}.xlsx"

    return send_file(
        workbook_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )