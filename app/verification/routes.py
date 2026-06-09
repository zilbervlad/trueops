from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    send_file,
    current_app,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.auth.routes import login_required, role_required
from app.extensions import db
from app.services.tenant import scoped_get_or_404
from app.models import (
    VerificationTemplateField,
    VerificationReport,
    VerificationReportValue,
    Store,
    User,
)
from app.services.email_service import send_email


verification_bp = Blueprint("verification", __name__, url_prefix="/verification")

def is_admin_like():
    return bool(session.get("is_platform_admin")) or session.get("user_role") in ["admin", "platform_admin"]




CORE_VERIFICATION_KEYS = {
    "bad_orders",
    "suspicious_activity",
    "csr_program",
    "dumpster_check",
}


def today_et():
    from zoneinfo import ZoneInfo

    return datetime.now(ZoneInfo("America/New_York")).date()


def get_active_company_id():
    """
    Returns the company currently being viewed.

    This supports multiple possible session keys because different parts of
    TrueOps have used slightly different names while the multi-company system
    was being built.
    """
    for key in ("selected_company_id", "active_company_id", "current_company_id", "company_id"):
        value = session.get(key)
        if value:
            try:
                return int(value)
            except (TypeError, ValueError):
                pass

    user_id = session.get("user_id")
    if user_id:
        user = db.session.get(User, user_id)
        if user and getattr(user, "company_id", None):
            return user.company_id

    return None


def verification_template_query(include_inactive=False):
    company_id = get_active_company_id()

    query = VerificationTemplateField.query

    if company_id and hasattr(VerificationTemplateField, "company_id"):
        query = query.filter(VerificationTemplateField.company_id == company_id)

    if not include_inactive:
        query = query.filter(VerificationTemplateField.is_active == True)

    return query


def ensure_company_verification_template(company_id):
    """
    If a company has no verification template yet, clone an existing/default template
    so each company can edit its own verification fields independently.
    """
    if not company_id:
        return

    existing_count = VerificationTemplateField.query.filter_by(company_id=company_id).count()
    if existing_count > 0:
        return

    source_items = VerificationTemplateField.query.filter(
        VerificationTemplateField.company_id.isnot(None),
        VerificationTemplateField.company_id != company_id,
    ).order_by(
        VerificationTemplateField.sort_order.asc(),
        VerificationTemplateField.id.asc(),
    ).all()

    if not source_items:
        source_items = VerificationTemplateField.query.filter(
            VerificationTemplateField.company_id.is_(None)
        ).order_by(
            VerificationTemplateField.sort_order.asc(),
            VerificationTemplateField.id.asc(),
        ).all()

    if not source_items:
        defaults = [
            ("bad_orders", "Bad order / cancel log system in place?", "textarea", 1),
            ("suspicious_activity", "Anyone identified for suspicious activity / callbacks made?", "textarea", 2),
            ("csr_program", "Is CSR development program in use?", "textarea", 3),
            ("dumpster_check", "Check dumpsters for waste - what did you see?", "textarea", 4),
        ]

        for field_key, field_label, field_type, sort_order in defaults:
            db.session.add(
                VerificationTemplateField(
                    company_id=company_id,
                    field_key=field_key,
                    field_label=field_label,
                    field_type=field_type,
                    sort_order=sort_order,
                    is_active=True,
                )
            )

        db.session.commit()
        return

    for item in source_items:
        db.session.add(
            VerificationTemplateField(
                company_id=company_id,
                field_key=item.field_key,
                field_label=item.field_label,
                field_type=item.field_type,
                sort_order=item.sort_order,
                is_active=item.is_active,
            )
        )

    db.session.commit()


def scoped_store_query():
    """
    Base store query locked to the active company.

    This prevents stores from Polish Pie, TrueOps, or any other company from
    leaking into each other's Verification dashboards.
    """
    query = Store.query.filter_by(is_active=True)

    company_id = get_active_company_id()
    if company_id and hasattr(Store, "company_id"):
        query = query.filter(Store.company_id == company_id)

    return query


def get_scoped_stores():
    return scoped_store_query().order_by(Store.store_number.asc()).all()


def get_scoped_store_numbers():
    return [store.store_number for store in get_scoped_stores()]


def scoped_verification_report_query():
    """
    Base verification report query locked to the active company.

    If VerificationReport has company_id, use it directly.
    If not, scope reports by the store numbers that belong to the active company.
    """
    query = VerificationReport.query

    company_id = get_active_company_id()

    if company_id and hasattr(VerificationReport, "company_id"):
        return query.filter(VerificationReport.company_id == company_id)

    scoped_store_numbers = get_scoped_store_numbers()

    if not scoped_store_numbers:
        return query.filter(False)

    return query.filter(VerificationReport.store_number.in_(scoped_store_numbers))


def get_supervisor_stores():
    role = session.get("user_role")
    user_area = session.get("user_area")

    query = scoped_store_query()

    if role == "admin":
        return query.order_by(Store.store_number.asc()).all()

    if role == "supervisor":
        return query.filter(
            Store.area_name == user_area
        ).order_by(Store.store_number.asc()).all()

    return []


def ensure_default_template():
    defaults = [
        ("bad_orders", "Bad order / cancel log system in place?", "textarea"),
        ("suspicious_activity", "Anyone identified for suspicious activity / callbacks made?", "textarea"),
        ("csr_program", "Is CSR development program in use?", "textarea"),
        ("dumpster_check", "Check dumpsters for waste - what did you see?", "textarea"),
    ]

    company_id = get_active_company_id()
    ensure_company_verification_template(company_id)

    existing = {f.field_key: f for f in verification_template_query(include_inactive=True).all()}

    for i, (key, label, ftype) in enumerate(defaults, start=1):
        if key not in existing:
            db.session.add(
                VerificationTemplateField(
                    company_id=company_id,
                    field_key=key,
                    field_label=label,
                    field_type=ftype,
                    sort_order=i,
                    is_active=True,
                )
            )

    db.session.commit()


def get_dashboard_week_range():
    """
    Uses ?week_offset=0 for current week, -1 for last week, -2 for two weeks ago, etc.
    Also supports ?week_start=YYYY-MM-DD for direct linking if needed.
    """
    today = today_et()

    week_start_raw = (request.args.get("week_start") or "").strip()
    week_offset_raw = (request.args.get("week_offset") or "0").strip()

    if week_start_raw:
        try:
            selected_date = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
            week_start = selected_date - timedelta(days=selected_date.weekday())
        except ValueError:
            week_start = today - timedelta(days=today.weekday())
    else:
        try:
            week_offset = int(week_offset_raw)
        except ValueError:
            week_offset = 0

        current_week_start = today - timedelta(days=today.weekday())
        week_start = current_week_start + timedelta(weeks=week_offset)

    week_end = week_start + timedelta(days=6)
    return week_start, week_end


@verification_bp.route("/")
@login_required
@role_required("admin", "supervisor")
def index():
    if is_admin_like():
        return redirect(url_for("verification.dashboard"))

    return redirect(url_for("verification.new_report"))


@verification_bp.route("/dashboard")
@login_required
@role_required("admin")
def dashboard():
    week_start, week_end = get_dashboard_week_range()
    current_week_start = today_et() - timedelta(days=today_et().weekday())
    week_offset = (week_start - current_week_start).days // 7

    week_start_label = week_start.strftime("%m/%d")
    week_end_label = week_end.strftime("%m/%d")
    week_range_label = f"{week_start.strftime('%m/%d/%Y')} – {week_end.strftime('%m/%d/%Y')}"

    stores = get_scoped_stores()

    reports = scoped_verification_report_query().order_by(
        VerificationReport.report_date.desc(),
        VerificationReport.created_at.desc(),
    ).all()

    weekly_reports = [
        report for report in reports
        if report.report_date and week_start <= report.report_date <= week_end
    ]

    latest_by_store = {}
    for report in reports:
        if report.store_number not in latest_by_store:
            latest_by_store[report.store_number] = report

    submitted_this_week_store_numbers = {
        report.store_number
        for report in weekly_reports
    }

    total_stores = len(stores)
    submitted_count = len(submitted_this_week_store_numbers)
    missing_count = max(total_stores - submitted_count, 0)
    overall_compliance = round((submitted_count / total_stores) * 100, 1) if total_stores else 0.0

    stores_by_area = defaultdict(list)
    for store in stores:
        area_name = store.area_name or "Unassigned"
        stores_by_area[area_name].append(store)

    area_summary_rows = []
    areas_fully_complete = 0

    for area_name, area_stores in sorted(stores_by_area.items()):
        area_store_numbers = {store.store_number for store in area_stores}
        submitted_area_store_numbers = area_store_numbers & submitted_this_week_store_numbers

        store_count = len(area_stores)
        submitted_area_count = len(submitted_area_store_numbers)
        missing_area_count = max(store_count - submitted_area_count, 0)
        compliance = round((submitted_area_count / store_count) * 100, 1) if store_count else 0.0

        if missing_area_count == 0 and store_count > 0:
            areas_fully_complete += 1

        missing_store_numbers = sorted(list(area_store_numbers - submitted_this_week_store_numbers))

        area_summary_rows.append(
            {
                "area_name": area_name,
                "store_count": store_count,
                "submitted_count": submitted_area_count,
                "missing_count": missing_area_count,
                "missing_store_numbers": missing_store_numbers,
                "compliance": compliance,
            }
        )

    weekly_reports = sorted(
        weekly_reports,
        key=lambda report: (
            report.report_date or datetime.min.date(),
            report.created_at or datetime.min,
        ),
        reverse=True,
    )

    return render_template(
        "verification_dashboard.html",
        stores=stores,
        latest_by_store=latest_by_store,
        week_start=week_start,
        week_end=week_end,
        week_start_label=week_start_label,
        week_end_label=week_end_label,
        week_range_label=week_range_label,
        week_offset=week_offset,
        submitted_count=submitted_count,
        missing_count=missing_count,
        total_stores=total_stores,
        overall_compliance=overall_compliance,
        areas_fully_complete=areas_fully_complete,
        area_summary_rows=area_summary_rows,
        weekly_reports=weekly_reports,
    )


@verification_bp.route("/export-weekly")
@login_required
@role_required("admin")
def export_weekly_file():
    week_start, week_end = get_dashboard_week_range()

    stores = get_scoped_stores()

    reports = scoped_verification_report_query().filter(
        VerificationReport.report_date >= week_start,
        VerificationReport.report_date <= week_end,
    ).order_by(
        VerificationReport.report_date.asc(),
        VerificationReport.store_number.asc(),
        VerificationReport.created_at.asc(),
    ).all()

    submitted_store_numbers = {report.store_number for report in reports}

    stores_by_area = defaultdict(list)
    for store in stores:
        area_name = store.area_name or "Unassigned"
        stores_by_area[area_name].append(store)

    wb = Workbook()

    # Summary sheet
    ws1 = wb.active
    ws1.title = "Summary"

    ws1["A1"] = "Weekly Verification Summary"
    ws1["A1"].font = Font(size=14, bold=True)

    ws1["A3"] = "Week Start"
    ws1["B3"] = str(week_start)
    ws1["A4"] = "Week End"
    ws1["B4"] = str(week_end)
    ws1["A5"] = "Total Stores"
    ws1["B5"] = len(stores)
    ws1["A6"] = "Stores Submitted"
    ws1["B6"] = len(submitted_store_numbers)
    ws1["A7"] = "Stores Missing"
    ws1["B7"] = max(len(stores) - len(submitted_store_numbers), 0)
    ws1["A8"] = "Compliance %"
    ws1["B8"] = round((len(submitted_store_numbers) / len(stores)) * 100, 1) if stores else 0.0

    for col in ["A", "B"]:
        ws1.column_dimensions[col].width = 22

    # Area summary sheet
    ws2 = wb.create_sheet("Area Summary")
    headers = ["Area", "Store Count", "Submitted", "Missing", "Compliance %", "Missing Stores"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for area_name, area_stores in sorted(stores_by_area.items()):
        area_store_numbers = {store.store_number for store in area_stores}
        submitted_area_store_numbers = area_store_numbers & submitted_store_numbers
        missing_store_numbers = sorted(list(area_store_numbers - submitted_store_numbers))

        store_count = len(area_stores)
        submitted_count = len(submitted_area_store_numbers)
        missing_count = max(store_count - submitted_count, 0)
        compliance = round((submitted_count / store_count) * 100, 1) if store_count else 0.0

        ws2.cell(row=row_num, column=1, value=area_name)
        ws2.cell(row=row_num, column=2, value=store_count)
        ws2.cell(row=row_num, column=3, value=submitted_count)
        ws2.cell(row=row_num, column=4, value=missing_count)
        ws2.cell(row=row_num, column=5, value=compliance)
        ws2.cell(row=row_num, column=6, value=", ".join(missing_store_numbers))
        row_num += 1

    for i, width in enumerate([18, 14, 12, 12, 14, 40], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # Reports sheet
    ws3 = wb.create_sheet("Reports")
    headers = ["Report ID", "Store", "Date", "Submitted By", "Created At"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for report in reports:
        ws3.cell(row=row_num, column=1, value=report.id)
        ws3.cell(row=row_num, column=2, value=report.store_number)
        ws3.cell(row=row_num, column=3, value=str(report.report_date))
        ws3.cell(row=row_num, column=4, value=report.supervisor_name or "")
        ws3.cell(
            row=row_num,
            column=5,
            value=report.created_at.strftime("%Y-%m-%d %I:%M %p") if report.created_at else "",
        )
        row_num += 1

    for i, width in enumerate([12, 12, 14, 22, 24], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = width

    # Report details sheet
    ws4 = wb.create_sheet("Report Details")
    headers = ["Report ID", "Store", "Date", "Submitted By", "Question", "Response"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws4.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for report in reports:
        values = VerificationReportValue.query.filter_by(report_id=report.id).order_by(
            VerificationReportValue.sort_order.asc(),
            VerificationReportValue.id.asc(),
        ).all()

        if not values:
            ws4.cell(row=row_num, column=1, value=report.id)
            ws4.cell(row=row_num, column=2, value=report.store_number)
            ws4.cell(row=row_num, column=3, value=str(report.report_date))
            ws4.cell(row=row_num, column=4, value=report.supervisor_name or "")
            ws4.cell(row=row_num, column=5, value="No fields")
            ws4.cell(row=row_num, column=6, value="")
            row_num += 1
            continue

        for value in values:
            ws4.cell(row=row_num, column=1, value=report.id)
            ws4.cell(row=row_num, column=2, value=report.store_number)
            ws4.cell(row=row_num, column=3, value=str(report.report_date))
            ws4.cell(row=row_num, column=4, value=report.supervisor_name or "")
            ws4.cell(row=row_num, column=5, value=value.field_label or value.field_key)
            ws4.cell(row=row_num, column=6, value=value.value_text or "")
            row_num += 1

    for i, width in enumerate([12, 12, 14, 22, 42, 70], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"verification_week_{week_start}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@verification_bp.route("/report/<int:report_id>")
@login_required
@role_required("admin")
def view_report(report_id):
    report = scoped_verification_report_query().filter(
        VerificationReport.id == report_id
    ).first_or_404()

    values = VerificationReportValue.query.filter_by(report_id=report.id).order_by(
        VerificationReportValue.sort_order.asc(),
        VerificationReportValue.id.asc(),
    ).all()

    store = scoped_store_query().filter_by(store_number=report.store_number).first()

    if not store:
        flash("You do not have access to that verification report.", "error")
        return redirect(url_for("verification.dashboard"))

    return render_template(
        "verification_report_detail.html",
        report=report,
        values=values,
        store=store,
    )


@verification_bp.route("/new", methods=["GET", "POST"])
@login_required
@role_required("admin", "supervisor")
def new_report():
    stores = get_supervisor_stores()

    if not stores:
        flash("No stores available for verification.", "error")
        return redirect(url_for("dashboard.home"))

    fields = VerificationTemplateField.query.filter_by(is_active=True).order_by(
        VerificationTemplateField.sort_order.asc(),
        VerificationTemplateField.id.asc(),
    ).all()

    allowed_store_numbers = {store.store_number for store in stores}

    if request.method == "POST":
        store_number = (request.form.get("store_number") or "").strip()

        if store_number not in allowed_store_numbers:
            flash("Invalid store selection.", "error")
            return redirect(url_for("verification.new_report"))

        report = VerificationReport(
            store_number=store_number,
            supervisor_name=session.get("user_name"),
            created_by_user_id=session.get("user_id"),
        )

        company_id = get_active_company_id()
        if company_id and hasattr(VerificationReport, "company_id"):
            report.company_id = company_id

        db.session.add(report)
        db.session.flush()

        for field in fields:
            value = (request.form.get(field.field_key) or "").strip()

            db.session.add(
                VerificationReportValue(
                    report_id=report.id,
                    template_field_id=field.id,
                    field_key=field.field_key,
                    field_label=field.field_label,
                    sort_order=field.sort_order,
                    value_text=value,
                )
            )

        db.session.commit()

        try:
            body = f"Verification Report - Store {store_number}\n\n"
            body += f"Submitted by: {session.get('user_name') or 'Unknown'}\n"
            body += f"Submitted at: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}\n\n"

            for field in fields:
                val = (request.form.get(field.field_key) or "").strip()
                body += f"{field.field_label}:\n{val or '—'}\n\n"

            to_email = (
                current_app.config.get("MAIL_DEFAULT_SENDER")
                or current_app.config.get("MAIL_USERNAME")
                or ""
            ).strip()

            if not to_email:
                raise ValueError("Missing MAIL_DEFAULT_SENDER / MAIL_USERNAME in environment settings.")

            supervisor_email = None
            user_id = session.get("user_id")
            if user_id:
                submitting_user = db.session.get(User, user_id)
                if submitting_user:
                    supervisor_email = submitting_user.get_notification_email()

            admin_query = User.query.filter_by(is_active=True)

            company_id = get_active_company_id()
            if company_id and hasattr(User, "company_id"):
                admin_query = admin_query.filter(User.company_id == company_id)

            admin_users = admin_query.filter(
                User.role.in_(["admin", "platform_admin"])
            ).all()

            admin_emails = [
                user.get_notification_email()
                for user in admin_users
                if user.get_notification_email()
            ]

            cc_list = []

            if supervisor_email:
                cc_list.append(supervisor_email)

            for email in admin_emails:
                if email and email not in cc_list and email != to_email:
                    cc_list.append(email)

            send_email(
                to_email=to_email,
                subject=f"Verification - Store {store_number}",
                body=body,
                cc_emails=cc_list if cc_list else None,
            )

        except Exception as e:
            print("Email failed:", e)

        flash("Verification submitted.", "success")
        return redirect(url_for("dashboard.home"))

    return render_template(
        "verification_form.html",
        stores=stores,
        fields=fields,
    )


@verification_bp.route("/admin", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin():
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "create":
            field_key = (request.form.get("field_key") or "").strip()
            field_label = (request.form.get("field_label") or "").strip()
            field_type = (request.form.get("field_type") or "textarea").strip()
            sort_order_raw = (request.form.get("sort_order") or "999").strip()

            if not field_key or not field_label:
                flash("Field key and label are required.", "error")
                return redirect(url_for("verification.admin"))

            try:
                sort_order = int(sort_order_raw)
            except ValueError:
                flash("Sort order must be a number.", "error")
                return redirect(url_for("verification.admin"))

            existing = verification_template_query(include_inactive=True).filter_by(field_key=field_key).first()
            if existing:
                flash("That field key already exists.", "error")
                return redirect(url_for("verification.admin"))

            db.session.add(
                VerificationTemplateField(
                    company_id=get_active_company_id(),
                    field_key=field_key,
                    field_label=field_label,
                    field_type=field_type,
                    sort_order=sort_order,
                    is_active=True,
                )
            )
            db.session.commit()
            flash("Verification field created.", "success")
            return redirect(url_for("verification.admin"))

        if action == "delete":
            field_id = (request.form.get("field_id") or "").strip()
            field = verification_template_query(include_inactive=True).filter_by(id=field_id).first()

            if not field:
                flash("Field not found.", "error")
                return redirect(url_for("verification.admin"))

            db.session.delete(field)
            db.session.commit()
            flash("Verification field deleted.", "success")
            return redirect(url_for("verification.admin"))

        if action == "update":
            field_id = (request.form.get("field_id") or "").strip()
            field = verification_template_query(include_inactive=True).filter_by(id=field_id).first()

            if not field:
                flash("Field not found.", "error")
                return redirect(url_for("verification.admin"))

            field.field_key = (request.form.get("field_key") or "").strip()
            field.field_label = (request.form.get("field_label") or "").strip()
            field.field_type = (request.form.get("field_type") or "textarea").strip()

            try:
                field.sort_order = int((request.form.get("sort_order") or "999").strip())
            except ValueError:
                flash("Sort order must be a number.", "error")
                return redirect(url_for("verification.admin"))

            field.is_active = "is_active" in request.form

            duplicate = VerificationTemplateField.query.filter(
                VerificationTemplateField.field_key == field.field_key,
                VerificationTemplateField.id != field.id,
            ).first()

            if duplicate:
                flash("That field key already exists.", "error")
                return redirect(url_for("verification.admin"))

            if not field.field_key or not field.field_label:
                flash("Field key and label are required.", "error")
                return redirect(url_for("verification.admin"))

            db.session.commit()
            flash("Verification field updated.", "success")
            return redirect(url_for("verification.admin"))

    fields = VerificationTemplateField.query.order_by(
        VerificationTemplateField.sort_order.asc(),
        VerificationTemplateField.id.asc(),
    ).all()

    return render_template(
        "verification_admin.html",
        fields=fields,
        core_field_keys=CORE_VERIFICATION_KEYS,
    )